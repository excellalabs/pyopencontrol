from collections import OrderedDict
from openpyxl import load_workbook
import os.path
import wget
from .control import Control

DEFAULT_FILENAME = 'FedRAMP-Moderate-HHH-Baseline-Controls-2016-05-18.xlsx'
DEFAULT_INPUT_XLSX = 'https://s3.amazonaws.com/sitesusa/wp-content/uploads/sites/482/2016/07/{}'.format(DEFAULT_FILENAME)

def parse(filename=DEFAULT_FILENAME, url=DEFAULT_INPUT_XLSX, inject_cm3_2=True):
    _find_or_get_file(filename, url)
    controls = _get_controls()
    if inject_cm3_2:
        cm3_2 = get_cm3_2()
        # TODO improve sorted algorithm
        for control in controls:
            if control.sortid > cm3_2.sortid:
                idx = controls.index(control)
                controls.insert(idx, cm3_2)
                break
    return OrderedDict([(ctl.id, ctl) for ctl in controls])


def _find_or_get_file(filename, url):
    path = './{}'.format(filename)
    if os.path.isfile(path):
        print('Found existing file: {}'.format(path))
    else:
        print('Could not find file: {}'.format(path))
        print('Downloading...')
        wget.download(url, out=filename)

def _get_controls():

    controls = []

    workbook = load_workbook(DEFAULT_FILENAME)
    if workbook.active._cells[(3, 2)].value != 'AC-01':
        raise Exception("Unrecognized format, expected cell B3 to have value 'AC-01'")
    for row in workbook.active.iter_rows(min_row=3):
        # 0 Count
        # 1 SORT ID
        # 2 Family
        # 3 ID
        # 4 Control Name
        # 5 "NIST Control Description (From NIST SP 800-53r4 1/23/15)"
        # 6 "Moderate 325"
        # 7 "Moderate FedRAMP-Defined Assignment / Selection Parameters (Numbering matches SSP)"
        # 8 LM Additional FedRAMP Requirements and Guidance
        # 9 Parameter

        sortid = row[1].value
        control = row[3].value
        name = row[4].value
        text = row[5].value
        if None in [sortid, control, name, text]:
            continue
        new_control = Control(sortid=sortid, id=control, name=name, text=text)
        controls.append(new_control)
    return controls

def get_cm3_2():
    # CM-3 (2) is missing from the 2016-05-18 doc
    sortid = 'CM-03 (02)'
    family = 'CM'
    control = 'CM-3 (2)'
    name = 'CONFIGURATION CHANGE CONTROL TEST / VALIDATE / DOCUMENT CHANGES'
    text = '''The organization tests, validates, and documents changes to the information system before implementing the changes on the operational system.

Supplemental Guidance: Changes to information systems include modifications to hardware, softwaee, or firmware components and configuration settings defined in CM-6. Organizations ensure that testing does not interfere with information system operations. Individuals/groups conducting tests understand organizational security policies and procedures, information system security policies and procedures, and the specific health, safety, and environmental risks associated with particular facilities/processes. Operational systems may need to be taken off-line, or replicated to the extent feasible, before testing can be conducted. If information systems must be taken off-line for testing, the tests are scheduled to occur during planned system outages whenever possible. If testing cannot be conducted on operational systems, organizations employ compensating controls (e.g., testing on replicated systems).'''
    return Control(sortid=sortid, id=control, name=name, text=text)
